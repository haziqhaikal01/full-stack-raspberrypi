# # frequency field (VERSION 5.1)
# from flask import Flask, jsonify, request, send_file
# from flask_cors import CORS
# from urllib.parse import unquote
# import os
# import ssl
# import certifi
# import pandas as pd
# import numpy as np
# from influxdb_client_3 import InfluxDBClient3
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import Font
# from openpyxl import load_workbook

# app = Flask(__name__)
# CORS(app)

# INFLUXDB_HOST = "https://eu-central-1-1.aws.cloud2.influxdata.com"
# INFLUXDB_TOKEN = os.getenv("INFLUXDB_TOKEN")
# INFLUXDB_DATABASE = "FrequencyMeas"

# if not INFLUXDB_TOKEN:
#     raise ValueError("❌ INFLUXDB_TOKEN is not set!")

# os.environ["GRPC_DEFAULT_SSL_ROOTS_FILE_PATH"] = certifi.where()
# ssl_context = ssl.create_default_context(cafile=certifi.where())

# LOCATION_MAPPING = {
#     "manchester": "M13 9PL",
#     "scotland": "S23 9RW",
# }

# @app.route("/frequency/<location>", methods=["GET"])
# def get_frequency(location):
#     decoded_location = unquote(location).strip()
#     minutes = request.args.get("minutes", default=1, type=int)

#     field = "frequency"  # 👈 CHANGE THIS FIELD as needed (e.g., "voltage", "current")

#     if decoded_location not in LOCATION_MAPPING.values():
#         return jsonify({"error": "Invalid location"}), 400

#     try:
#         client = InfluxDBClient3(
#             host=INFLUXDB_HOST,
#             token=INFLUXDB_TOKEN,
#             database=INFLUXDB_DATABASE,
#             ssl_context=ssl_context
#         )

#         query = f"""
#         SELECT time, {field}
#         FROM "FrequencyTest"
#         WHERE "location" = '{decoded_location}'
#         AND time >= now() - interval '{minutes} minutes'
#         ORDER BY time ASC
#         """

#         result = client.query(query)
#         client.close()

#         df = result.to_pandas()

#         if df.empty or field not in df.columns:
#             return jsonify([]), 404

#         df = df.dropna(subset=[field])
#         df[field] = df[field].round(3)
#         df["time"] = pd.to_datetime(df["time"])
#         df["time"] = df["time"].dt.strftime("%Y-%m-%dT%H:%M:%S.%f").str[:-3] + "Z"

#         return jsonify(df.to_dict(orient="records"))

#     except Exception as e:
#         print(f"❌ Error processing request: {e}")
#         return jsonify({"error": str(e)}), 500


# @app.route("/fetch-data", methods=["GET"])
# def fetch_data():
#     try:
#         field = "frequency"  # 👈 CHANGE THIS FIELD as needed (e.g., "voltage", "current")

#         start = request.args.get("start")
#         end = request.args.get("end")
#         location = request.args.get("location")

#         if not start or not end or not location:
#             return jsonify({"error": "Missing parameters"}), 400

#         if location.lower() not in LOCATION_MAPPING:
#             return jsonify({"error": "Invalid location"}), 400

#         influx_location = LOCATION_MAPPING[location.lower()]

#         client = InfluxDBClient3(
#             host=INFLUXDB_HOST,
#             token=INFLUXDB_TOKEN,
#             database=INFLUXDB_DATABASE,
#             ssl_context=ssl_context
#         )

#         query = f"""
#         SELECT time, {field}
#         FROM "FrequencyTest"
#         WHERE "location" = '{decoded_location}'
#         AND time BETWEEN '{start}' AND '{end}'
#         ORDER BY time DESC
#         """

#         result = client.query(query)
#         client.close()

#         df = result.to_pandas()

#         if df.empty or field not in df.columns:
#             return jsonify({"error": "No data found"}), 404

#         df[field] = df[field].astype(float).map(lambda x: f"{x:.3f}")
#         df["time"] = pd.to_datetime(df["time"]).dt.strftime("%Y-%m-%d %H:%M:%S.%f").str[:-3]

#         file_path = f"output_{location}_{field}.xlsx"
#         with pd.ExcelWriter(file_path, engine="openpyxl", datetime_format="yyyy-mm-dd hh:mm:ss.000") as writer:
#             df.to_excel(writer, index=False, sheet_name="Data")

#         wb = load_workbook(file_path)
#         ws = wb["Data"]

#         for cell in ws[1]:
#             cell.font = Font(bold=True)

#         for col in ws.columns:
#             max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
#             col_letter = get_column_letter(col[0].column)
#             ws.column_dimensions[col_letter].width = max_len + 2

#         wb.save(file_path)

#         return send_file(
#             file_path,
#             as_attachment=True,
#             download_name=f"influx-data-{location}-{field}.xlsx",
#             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )

#     except Exception as e:
#         print(f"❌ Error: {e}")
#         return jsonify({"error": str(e)}), 500


# if __name__ == "__main__":
#     app.run(host="0.0.0.0", port=5000, debug=True)




# Frequency field (VERSION 5.2)
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from urllib.parse import unquote
import os
import ssl
import certifi
import pandas as pd
import numpy as np
from influxdb_client_3 import InfluxDBClient3
from influxdb_client import InfluxDBClient
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl import load_workbook
from pyarrow import Table
import os
from dotenv import load_dotenv



app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})

INFLUXDB_HOST = "https://eu-central-1-1.aws.cloud2.influxdata.com"
INFLUXDB_HOST_LOCAL = "http://localhost:8086"


INFLUXDB_TOKEN = os.getenv("INFLUXDB_TOKEN", "").strip().replace('"', '').replace("'", '')
INFLUXDB_TOKEN_LOCAL = "6TyaYt-PGfYlcR4KVLWERtpFxuu5SA2NIufL2YZjhj-4RJ83wEQQ10T46YRDsFp1ackPJfUj2HXVlh6NTOG7GA=="

INFLUXDB_DATABASE = "FrequencyMeas"
INFLUXDB_DATABASE_LOCAL = "RaspiDevice"
INFLUXDB_ORG = "51aac14ba4cafd71"

if not INFLUXDB_TOKEN:
    raise ValueError("❌ INFLUXDB_TOKEN is not set!")

os.environ["GRPC_DEFAULT_SSL_ROOTS_FILE_PATH"] = certifi.where()
ssl_context = ssl.create_default_context(cafile=certifi.where())

LOCATION_MAPPING = {
    "manchester": "M13 9RW",
    "scotland": "S23 9RW",
}


@app.route("/frequency/<location>", methods=["GET"])
def get_frequency(location):
    print("📥 Request received to fetch Frequency field")

    minutes = request.args.get("minutes", default=1, type=int)
    field = "Frequency"  # Keep capital F

    location_key = location.lower().strip()
    if location_key not in LOCATION_MAPPING:
        return jsonify({"error": "Invalid location"}), 400

    influx_location = LOCATION_MAPPING[location_key]

    try:
        client = InfluxDBClient3(
            host=INFLUXDB_HOST,
            token=INFLUXDB_TOKEN,
            database=INFLUXDB_DATABASE,
            ssl_context=ssl_context
        )

        query = f"""
        SELECT time, "Frequency"
        FROM "FrequencyTest"
        WHERE "location" = '{influx_location}'
        AND "Frequency" IS NOT NULL
        AND time >= now() - interval '{minutes} minutes'
        ORDER BY time ASC
        """

        print("🧠 Final query:\n", query)

        table = client.query(query)  # PyArrow Table
        client.close()

        time_col = table.column("time").to_pylist()
        freq_col = table.column(field).to_pylist()

        rows = []
        for t, v in zip(time_col, freq_col):
            if v is not None:
                rows.append({
                    "time": pd.to_datetime(t).strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z",
                    "frequency": round(float(v), 3)
                })

        if not rows:
            return jsonify([]), 404

        print("✅ Returned rows:", rows[:3])
        return jsonify(rows)

    except Exception as e:
        print(f"❌ Error querying InfluxDB: {e}")
        return jsonify({"error": str(e)}), 500



@app.route("/fetch-data", methods=["GET"])
def fetch_data():
    start = request.args.get("start")
    end = request.args.get("end")
    location = request.args.get("location")

    if not start or not end or not location:
        print("❌ Missing parameters in request")
        return jsonify({"error": "Start date, end date, and location are required"}), 400

    try:
        location_key = location.strip().lower()
        if location_key not in LOCATION_MAPPING:
            print(f"❌ Invalid location requested: {location}")
            return jsonify({"error": "Invalid location"}), 400

        influx_location = LOCATION_MAPPING[location_key]

        print(f"📡 Fetching historical data for {location_key} ({influx_location})")
        print(f"   ➤ Start: {start}")
        print(f"   ➤ End:   {end}")
        print("🔍 CONFIG CHECK:")
        print(f"   ➤ Host:   {INFLUXDB_HOST_LOCAL}")
        print(f"   ➤ Org:    {INFLUXDB_ORG}")
        print(f"   ➤ Bucket: {INFLUXDB_DATABASE_LOCAL}")
        print(f"   ➤ Token:  {INFLUXDB_TOKEN_LOCAL[:5]}...")

        # Flux query
        flux_query = f'''
        from(bucket: "{INFLUXDB_DATABASE_LOCAL}")
          |> range(start: time(v: "{start}"), stop: time(v: "{end}"))
          |> filter(fn: (r) => r._measurement == "FrequencyTest")
          |> filter(fn: (r) => r.location == "{influx_location}")
          |> filter(fn: (r) => r._field == "Frequency")
          |> sort(columns: ["_time"], desc: true)
        '''

        client = InfluxDBClient(
            url=INFLUXDB_HOST_LOCAL,
            token=INFLUXDB_TOKEN_LOCAL,
            org=INFLUXDB_ORG
        )

        query_api = client.query_api()
        try:
            df = query_api.query_data_frame(flux_query)
            print(f"✅ Query returned {df.shape[0]} rows")
        except Exception as query_err:
            print(f"❌ Error during query execution: {query_err}")
            return jsonify({"error": f"Query failed: {query_err}"}), 500

        if df.empty:
            print("⚠️ No data found for query.")
            return jsonify({"error": "No data found"}), 404

        # Format DataFrame
        df = df.rename(columns={"_time": "time", "_value": "frequency"})
        df["frequency"] = df["frequency"].astype(float).map(lambda x: f"{x:.3f}")
        df["time"] = pd.to_datetime(df["time"]).dt.strftime("%Y-%m-%d %H:%M:%S.%f").str[:-3]

        file_path = f"output_{location}.xlsx"
        with pd.ExcelWriter(file_path, engine="openpyxl", datetime_format="yyyy-mm-dd hh:mm:ss.000") as writer:
            df[["time", "frequency"]].to_excel(writer, index=False, sheet_name="Data")

        wb = load_workbook(file_path)
        ws = wb["Data"]

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_len + 2

        wb.save(file_path)

        return send_file(
            file_path,
            as_attachment=True,
            download_name=f"influx-data-{location}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print(f"❌ Error fetching data: {e}")
        return jsonify({"error": str(e)}), 500





if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5050, debug=True)